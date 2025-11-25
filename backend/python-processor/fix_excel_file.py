#!/usr/bin/env python3
"""
Fix the Excel file by removing problematic view settings and saving a clean version
"""

import os
import sys
from openpyxl import load_workbook
import pandas as pd

def fix_excel_file():
    """Fix the Excel file by removing view settings that cause issues"""
    
    input_path = "/Users/diegoalejandroparraruiz/Documents/Aova/Embler_Stats/ventas cuautemoc 05 de mayo al 05 agosto.xlsx"
    output_path = "/Users/diegoalejandroparraruiz/Documents/Aova/Embler_Stats/ventas cuautemoc 05 de mayo al 05 agosto_fixed.xlsx"
    
    print("=" * 60)
    print("FIXING EXCEL FILE")
    print("=" * 60)
    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")
    print()
    
    if not os.path.exists(input_path):
        print(f"❌ File not found: {input_path}")
        return
    
    try:
        print("🔧 Attempting to fix Excel file...")
        
        # Method 1: Try to load and save with minimal settings
        try:
            # Load workbook without reading formulas or view settings
            wb = load_workbook(input_path, read_only=True, data_only=True)
            
            # Create a new workbook and copy data
            from openpyxl import Workbook
            new_wb = Workbook()
            new_ws = new_wb.active
            
            # Copy data from original worksheet
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                new_ws.append(row)
            
            # Save the cleaned version
            new_wb.save(output_path)
            print(f"✅ Successfully created fixed file: {output_path}")
            
            # Test if the fixed file can be read by pandas
            test_df = pd.read_excel(output_path)
            print(f"✅ Fixed file can be read by pandas: {test_df.shape}")
            print("   Sample data:")
            print(test_df.head(3))
            
            return True
            
        except Exception as method1_error:
            print(f"❌ Method 1 failed: {method1_error}")
            
            # Method 2: Try using xlwings or other libraries
            try:
                import xlwings as xw
                print("🔧 Trying xlwings method...")
                
                app = xw.App(visible=False)
                wb = app.books.open(input_path)
                wb.save(output_path)
                wb.close()
                app.quit()
                
                # Test the result
                test_df = pd.read_excel(output_path)
                print(f"✅ Fixed file with xlwings: {test_df.shape}")
                return True
                
            except ImportError:
                print("⚠️  xlwings not available, trying manual data extraction...")
                
                # Method 3: Manual data extraction using different approach
                try:
                    import zipfile
                    import xml.etree.ElementTree as ET
                    
                    print("🔧 Trying manual XML extraction...")
                    
                    # Excel files are ZIP archives, extract the data manually
                    with zipfile.ZipFile(input_path, 'r') as zip_file:
                        # Read the worksheet data
                        sheet_xml = zip_file.read('xl/worksheets/sheet1.xml')
                        root = ET.fromstring(sheet_xml)
                        
                        # This is a complex approach, let's try a simpler one
                        print("⚠️  Manual extraction is complex, trying pandas with different parameters...")
                        
                        # Try reading with different pandas parameters
                        test_df = pd.read_excel(input_path, engine='openpyxl', header=None)
                        print("❌ This shouldn't work but let's see...")
                        
                except Exception as method3_error:
                    print(f"❌ Method 3 failed: {method3_error}")
                    return False
            
            except Exception as method2_error:
                print(f"❌ Method 2 failed: {method2_error}")
                return False
    
    except Exception as e:
        print(f"❌ General error: {e}")
        return False

def suggest_manual_fix():
    """Suggest manual steps to fix the file"""
    print("\n" + "=" * 60)
    print("MANUAL FIX SUGGESTION")
    print("=" * 60)
    print("The Excel file has view settings that are incompatible with the current")
    print("version of openpyxl. Here are some solutions:")
    print()
    print("1. Open the file in Excel and save it as a new file:")
    print("   - Open: ventas cuautemoc 05 de mayo al 05 agosto.xlsx")
    print("   - Save As: ventas cuautemoc 05 de mayo al 05 agosto_fixed.xlsx")
    print("   - Make sure to save as 'Excel Workbook (.xlsx)'")
    print()
    print("2. Or try opening in Google Sheets and downloading as Excel")
    print()
    print("3. Or use a different Excel viewer to resave the file")
    print()
    print("Once you have a fixed version, upload that instead.")

if __name__ == "__main__":
    success = fix_excel_file()
    if not success:
        suggest_manual_fix()
