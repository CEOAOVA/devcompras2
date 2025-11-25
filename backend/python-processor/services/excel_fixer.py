"""
Excel File Fixer - Handles problematic Excel files with view settings issues
"""

import os
import tempfile
import shutil
from typing import Optional
import logging

logger = logging.getLogger(__name__)

class ExcelFileFixer:
    """Fixes problematic Excel files that have view settings issues"""
    def __init__(self):
        self.temp_files = []
    
    def fix_excel_file(self, file_path: str) -> Optional[str]:
        """
        Fix an Excel file that has view settings issues (WEB-SAFE VERSION)
        Returns path to fixed file, or None if fixing failed
        """
        try:
            # Try to read the file normally first
            import pandas as pd
            try:
                pd.read_excel(file_path, nrows=1)
                # If this works, no fixing needed
                return file_path
            except Exception as e:
                error_msg = str(e)
                if "WindowWidth" in error_msg:
                    logger.warning(f"Excel file has view settings corruption: {file_path}")
                    # This type of corruption requires manual fixing by user
                    # Return None to trigger user-friendly error message
                    return None
                elif "BookView" in error_msg:
                    logger.warning(f"Excel file has view corruption: {file_path}")
                    # Try web-safe fixing methods
                    pass
                else:
                    # Different error, can't fix
                    logger.error(f"Excel file has unrecognized error: {error_msg}")
                    return None
            
            # Create a temporary fixed file for web-safe methods only
            temp_dir = tempfile.gettempdir()
            base_name = os.path.basename(file_path)
            name, ext = os.path.splitext(base_name)
            fixed_path = os.path.join(temp_dir, f"{name}_websafe_fixed{ext}")
            
            # Method 1: Try openpyxl with various parameters (WEB-SAFE)
            if self._fix_with_openpyxl(file_path, fixed_path):
                self.temp_files.append(fixed_path)
                return fixed_path
            
            # Method 2: Try pandas with different engines (WEB-SAFE)
            if self._fix_with_pandas_engines(file_path, fixed_path):
                self.temp_files.append(fixed_path)
                return fixed_path
            
            # If all web-safe methods fail, return None to trigger user guidance
            logger.warning(f"Could not fix Excel file with web-safe methods: {file_path}")
            return None
            
        except Exception as e:
            logger.error(f"Error in Excel file fixer: {e}")
            return None
    
    def _fix_with_xlwings(self, input_path: str, output_path: str) -> bool:
        """Fix using xlwings (NOT SAFE FOR WEB DEPLOYMENT - DISABLED)"""
        # DISABLED FOR WEB DEPLOYMENT - xlwings requires Excel and user permissions
        return False
    
    def _fix_with_openpyxl(self, input_path: str, output_path: str) -> bool:
        """Fix using openpyxl read-only mode (WEB-SAFE)"""
        try:
            from openpyxl import load_workbook, Workbook
            logger.info("Attempting to fix Excel file with openpyxl (web-safe method)...")
            
            # Method 1: Try read-only with data_only
            try:
                wb = load_workbook(input_path, read_only=True, data_only=True)
                new_wb = Workbook()
                new_ws = new_wb.active
                
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    new_ws.append(row)
                
                new_wb.save(output_path)
                
                # Test if the fixed file works
                import pandas as pd
                pd.read_excel(output_path, nrows=1)
                logger.info("Successfully fixed Excel file with openpyxl read-only method")
                return True
                
            except Exception as method1_error:
                logger.debug(f"openpyxl read-only method failed: {method1_error}")
                
                # Method 2: Try with keep_vba=False to strip problematic elements
                try:
                    wb = load_workbook(input_path, keep_vba=False, data_only=True)
                    new_wb = Workbook()
                    new_ws = new_wb.active
                    
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        new_ws.append(row)
                    
                    new_wb.save(output_path)
                    
                    # Test if the fixed file works
                    import pandas as pd
                    pd.read_excel(output_path, nrows=1)
                    logger.info("Successfully fixed Excel file with openpyxl keep_vba=False method")
                    return True
                    
                except Exception as method2_error:
                    logger.debug(f"openpyxl keep_vba=False method failed: {method2_error}")
                    return False
            
        except Exception as e:
            logger.debug(f"openpyxl method completely failed: {e}")
            return False
    
    def _fix_with_pandas_engines(self, input_path: str, output_path: str) -> bool:
        """Fix using pandas with different engines (WEB-SAFE)"""
        try:
            import pandas as pd
            logger.info("Attempting to fix Excel file with pandas engines...")
            
            # Try different pandas engines
            engines = ['openpyxl', 'xlrd', 'calamine']
            
            for engine in engines:
                try:
                    logger.debug(f"Trying pandas with {engine} engine...")
                    # Read with specific engine
                    df = pd.read_excel(input_path, engine=engine, header=None)
                    
                    # Write to new file with openpyxl (most compatible)
                    df.to_excel(output_path, index=False, header=False, engine='openpyxl')
                    
                    # Test if the fixed file works
                    pd.read_excel(output_path, nrows=1)
                    logger.info(f"Successfully fixed Excel file with pandas {engine} engine")
                    return True
                    
                except Exception as engine_error:
                    logger.debug(f"pandas {engine} engine failed: {engine_error}")
                    continue
            
            return False
            
        except Exception as e:
            logger.debug(f"pandas engines method failed: {e}")
            return False
    
    def _fix_with_manual_extraction(self, input_path: str, output_path: str) -> bool:
        """Fix using manual data extraction (last resort)"""
        try:
            # This would be a complex implementation
            # For now, just return False
            logger.debug("Manual extraction not implemented yet")
            return False
            
        except Exception as e:
            logger.debug(f"Manual extraction failed: {e}")
            return False
    
    def cleanup(self):
        """Clean up temporary files"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                logger.warning(f"Could not delete temp file {temp_file}: {e}")
        self.temp_files.clear()
    
    def __del__(self):
        """Cleanup when object is destroyed"""
        self.cleanup()

# Convenience function
def fix_excel_file_if_needed(file_path: str) -> Optional[str]:
    """
    Fix an Excel file if it has view settings issues
    Returns path to fixed file (may be the same as input if no fixing needed)
    """
    fixer = ExcelFileFixer()
    try:
        return fixer.fix_excel_file(file_path)
    finally:
        # Note: Don't cleanup here as the caller needs the fixed file
        pass
