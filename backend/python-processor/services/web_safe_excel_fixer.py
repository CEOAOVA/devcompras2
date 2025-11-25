"""
Web-Safe Excel File Fixer - Production Ready

This module provides web deployment-safe methods to fix corrupted Excel files
without requiring desktop applications or user permissions.

Designed specifically for handling files with BookView/WindowWidth corruption
like the Cuautemoc file issue.
"""

import os
import tempfile
import shutil
import logging
from typing import Optional, Tuple
import io
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

logger = logging.getLogger(__name__)

class WebSafeExcelFixer:
    """
    Production-ready Excel fixer that works in web deployment environments.
    No desktop dependencies, no user permissions required.
    """
    
    def __init__(self):
        self.temp_files = []
        self.supported_methods = [
            'pandas_engines',
            'openpyxl_data_only', 
            'xml_reconstruction',
            'csv_fallback'
        ]
    
    def fix_excel_file(self, file_path: str) -> Optional[str]:
        """
        Fix Excel file using web-safe methods only.
        Returns path to fixed file or None if unfixable.
        """
        try:
            # First, test if file needs fixing
            if self._test_file_readability(file_path):
                logger.info(f"File {file_path} is already readable, no fixing needed")
                return file_path
            
            logger.info(f"File {file_path} needs fixing, trying web-safe methods...")
            
            # Create temp file for fixed version
            temp_dir = tempfile.gettempdir()
            base_name = Path(file_path).stem
            fixed_path = os.path.join(temp_dir, f"{base_name}_websafe_fixed.xlsx")
            
            # Try each method in order of reliability
            for method_name in self.supported_methods:
                logger.info(f"Trying method: {method_name}")
                method = getattr(self, f'_fix_with_{method_name}')
                
                if method(file_path, fixed_path):
                    if self._test_file_readability(fixed_path):
                        logger.info(f"Successfully fixed file with method: {method_name}")
                        self.temp_files.append(fixed_path)
                        return fixed_path
                    else:
                        logger.warning(f"Method {method_name} created file but it's not readable")
                        # Clean up failed attempt
                        if os.path.exists(fixed_path):
                            os.unlink(fixed_path)
            
            logger.error(f"All web-safe methods failed for file: {file_path}")
            return None
            
        except Exception as e:
            logger.error(f"Error in web-safe Excel fixer: {e}")
            return None
    
    def _test_file_readability(self, file_path: str) -> bool:
        """Test if an Excel file can be read by pandas"""
        try:
            import pandas as pd
            # Try to read just the first row to test readability
            df = pd.read_excel(file_path, nrows=1, engine='openpyxl')
            return True
        except Exception as e:
            logger.debug(f"File readability test failed: {e}")
            return False
    
    def _fix_with_pandas_engines(self, input_path: str, output_path: str) -> bool:
        """Try different pandas engines to read and rewrite the file"""
        try:
            import pandas as pd
            
            # List of engines to try (in order of preference)
            engines = ['openpyxl', 'xlrd']
            
            for engine in engines:
                try:
                    logger.debug(f"Trying pandas engine: {engine}")
                    
                    # Read with specific engine and various parameters
                    read_params = [
                        {'engine': engine},
                        {'engine': engine, 'header': None},
                        {'engine': engine, 'header': 0},
                    ]
                    
                    for params in read_params:
                        try:
                            df = pd.read_excel(input_path, **params)
                            
                            # Write to new file with openpyxl (most reliable for output)
                            df.to_excel(output_path, index=False, engine='openpyxl')
                            logger.info(f"Successfully processed with pandas {engine} engine")
                            return True
                            
                        except Exception as param_error:
                            logger.debug(f"Pandas {engine} with params {params} failed: {param_error}")
                            continue
                            
                except Exception as engine_error:
                    logger.debug(f"Pandas engine {engine} failed: {engine_error}")
                    continue
            
            return False
            
        except Exception as e:
            logger.debug(f"Pandas engines method failed: {e}")
            return False
    
    def _fix_with_openpyxl_data_only(self, input_path: str, output_path: str) -> bool:
        """Use openpyxl with data_only and read_only to bypass view issues"""
        try:
            from openpyxl import load_workbook, Workbook
            
            # Try multiple openpyxl configurations
            configs = [
                {'read_only': True, 'data_only': True},
                {'read_only': True, 'data_only': True, 'keep_vba': False},
                {'data_only': True, 'keep_vba': False},
                {'read_only': False, 'data_only': True, 'keep_vba': False}
            ]
            
            for config in configs:
                try:
                    logger.debug(f"Trying openpyxl config: {config}")
                    
                    # Load with specific configuration
                    wb = load_workbook(input_path, **config)
                    
                    # Create new clean workbook
                    new_wb = Workbook()
                    new_ws = new_wb.active
                    
                    # Copy data from original
                    ws = wb.active
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                        for col_idx, value in enumerate(row, 1):
                            new_ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Save clean version
                    new_wb.save(output_path)
                    logger.info(f"Successfully processed with openpyxl config: {config}")
                    return True
                    
                except Exception as config_error:
                    logger.debug(f"openpyxl config {config} failed: {config_error}")
                    continue
            
            return False
            
        except Exception as e:
            logger.debug(f"openpyxl data_only method failed: {e}")
            return False
    
    def _fix_with_xml_reconstruction(self, input_path: str, output_path: str) -> bool:
        """Reconstruct Excel file by extracting data from XML and rebuilding"""
        try:
            logger.debug("Attempting XML reconstruction method")
            
            # Excel files are ZIP archives
            with zipfile.ZipFile(input_path, 'r') as zip_file:
                # Try to extract worksheet data directly from XML
                try:
                    # Get the main worksheet XML
                    worksheet_xml = zip_file.read('xl/worksheets/sheet1.xml')
                    shared_strings_xml = None
                    
                    try:
                        shared_strings_xml = zip_file.read('xl/sharedStrings.xml')
                    except KeyError:
                        logger.debug("No shared strings found")
                    
                    # Parse the XML to extract cell data
                    data = self._extract_data_from_worksheet_xml(worksheet_xml, shared_strings_xml)
                    
                    if data:
                        # Create new Excel file with extracted data
                        import pandas as pd
                        df = pd.DataFrame(data)
                        df.to_excel(output_path, index=False, header=False, engine='openpyxl')
                        logger.info("Successfully reconstructed Excel file from XML")
                        return True
                    
                except Exception as xml_error:
                    logger.debug(f"XML extraction failed: {xml_error}")
                    return False
            
            return False
            
        except Exception as e:
            logger.debug(f"XML reconstruction method failed: {e}")
            return False
    
    def _extract_data_from_worksheet_xml(self, worksheet_xml: bytes, shared_strings_xml: bytes = None) -> list:
        """Extract cell data from worksheet XML"""
        try:
            # Parse shared strings if available
            shared_strings = []
            if shared_strings_xml:
                ss_root = ET.fromstring(shared_strings_xml)
                for si in ss_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                    t = si.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                    if t is not None:
                        shared_strings.append(t.text or '')
            
            # Parse worksheet
            ws_root = ET.fromstring(worksheet_xml)
            
            # Find all cells
            cells = {}
            for c in ws_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                ref = c.get('r')  # Cell reference like A1, B2, etc.
                if ref:
                    v = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                    t = c.get('t')  # Cell type
                    
                    if v is not None:
                        value = v.text
                        
                        # Handle different cell types
                        if t == 's' and value:  # Shared string
                            try:
                                idx = int(value)
                                if 0 <= idx < len(shared_strings):
                                    value = shared_strings[idx]
                            except (ValueError, IndexError):
                                pass
                        elif t == 'n':  # Number
                            try:
                                value = float(value) if '.' in value else int(value)
                            except ValueError:
                                pass
                        
                        cells[ref] = value
            
            # Convert to 2D array
            if not cells:
                return []
            
            # Find max row and column
            max_row = 0
            max_col = 0
            
            for ref in cells.keys():
                row, col = self._parse_cell_reference(ref)
                max_row = max(max_row, row)
                max_col = max(max_col, col)
            
            # Create 2D array
            data = []
            for r in range(1, max_row + 1):
                row_data = []
                for c in range(1, max_col + 1):
                    cell_ref = self._create_cell_reference(r, c)
                    value = cells.get(cell_ref, '')
                    row_data.append(value)
                data.append(row_data)
            
            return data
            
        except Exception as e:
            logger.debug(f"XML data extraction failed: {e}")
            return []
    
    def _parse_cell_reference(self, ref: str) -> Tuple[int, int]:
        """Parse cell reference like A1 into row, col numbers"""
        col_str = ''
        row_str = ''
        
        for char in ref:
            if char.isalpha():
                col_str += char
            else:
                row_str += char
        
        # Convert column letters to number
        col = 0
        for char in col_str:
            col = col * 26 + (ord(char.upper()) - ord('A') + 1)
        
        row = int(row_str)
        return row, col
    
    def _create_cell_reference(self, row: int, col: int) -> str:
        """Create cell reference like A1 from row, col numbers"""
        col_str = ''
        while col > 0:
            col -= 1
            col_str = chr(ord('A') + col % 26) + col_str
            col //= 26
        
        return f"{col_str}{row}"
    
    def _fix_with_csv_fallback(self, input_path: str, output_path: str) -> bool:
        """Last resort: try to read as CSV and convert to Excel"""
        try:
            import pandas as pd
            
            logger.debug("Attempting CSV fallback method")
            
            # Try to read as CSV with different separators
            separators = [',', ';', '\t', '|']
            
            for sep in separators:
                try:
                    # Try reading as CSV
                    df = pd.read_csv(input_path, sep=sep, encoding='utf-8', on_bad_lines='skip')
                    
                    if len(df.columns) > 1 and len(df) > 0:
                        # Save as Excel
                        df.to_excel(output_path, index=False, engine='openpyxl')
                        logger.info(f"Successfully converted from CSV with separator '{sep}'")
                        return True
                        
                except Exception as sep_error:
                    logger.debug(f"CSV separator '{sep}' failed: {sep_error}")
                    continue
            
            return False
            
        except Exception as e:
            logger.debug(f"CSV fallback method failed: {e}")
            return False
    
    def cleanup(self):
        """Clean up temporary files"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
                    logger.debug(f"Cleaned up temp file: {temp_file}")
            except Exception as e:
                logger.warning(f"Could not delete temp file {temp_file}: {e}")
        self.temp_files.clear()
    
    def __del__(self):
        """Cleanup when object is destroyed"""
        self.cleanup()

# Convenience function for single file fixing
def fix_excel_file_web_safe(file_path: str) -> Optional[str]:
    """
    Fix an Excel file using only web-safe methods.
    Returns path to fixed file or None if unfixable.
    """
    fixer = WebSafeExcelFixer()
    try:
        return fixer.fix_excel_file(file_path)
    except Exception as e:
        logger.error(f"Web-safe Excel fixing failed: {e}")
        return None
    # Note: Don't cleanup here as caller needs the fixed file

# Error handling for user guidance
class ExcelFixingError(Exception):
    """Custom exception for Excel fixing issues"""
    def __init__(self, message: str, user_action_required: bool = False, suggestions: list = None):
        self.message = message
        self.user_action_required = user_action_required
        self.suggestions = suggestions or []
        super().__init__(self.message)

def get_user_friendly_error_response(file_path: str, error: Exception) -> dict:
    """
    Generate user-friendly error response for Excel processing failures
    """
    error_str = str(error).lower()
    
    if 'windowwidth' in error_str or 'bookview' in error_str:
        return {
            'error_type': 'excel_view_corruption',
            'message': 'Your Excel file has view settings that are incompatible with our system.',
            'user_action_required': True,
            'suggestions': [
                'Open your Excel file in Microsoft Excel or Google Sheets',
                'Save it as a new .xlsx file (File → Save As → Excel Workbook)',
                'Upload the newly saved file',
                'Alternatively, copy the data to a new Excel file'
            ],
            'technical_details': f'File corruption type: {error_str}'
        }
    elif 'permission' in error_str or 'access' in error_str:
        return {
            'error_type': 'file_access',
            'message': 'Cannot access the Excel file due to permission restrictions.',
            'user_action_required': True,
            'suggestions': [
                'Make sure the file is not open in Excel',
                'Check file permissions',
                'Try copying the file to a different location and uploading again'
            ]
        }
    else:
        return {
            'error_type': 'general_excel_error',
            'message': 'Unable to process this Excel file.',
            'user_action_required': True,
            'suggestions': [
                'Verify the file is a valid Excel (.xlsx) file',
                'Try opening and re-saving the file in Excel',
                'Check if the file contains the expected data format'
            ],
            'technical_details': str(error)
        }
