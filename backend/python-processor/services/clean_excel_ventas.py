import pandas as pd
from . import get_sucursal as gs
import os

def excel_dataframization_ventas(path : str):

    # Read with error handling for problematic Excel files
    header_row = None
    try:
        raw_df = pd.read_excel(path)
    except Exception as e:
        # Handle Excel files with view settings issues (WindowWidth error)
        print(f"Warning: Standard Excel reading failed ({e}), trying alternative method...")
        try:
            # Try reading with xlrd engine (for older Excel files)
            raw_df = pd.read_excel(path, engine='xlrd')
        except:
            try:
                # Try reading with openpyxl but ignore view settings
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                raw_df = pd.DataFrame(data)
            except Exception as final_error:
                raise ValueError(f"Could not read Excel file {path}: {final_error}")
    
    # Find the header row    
    for i, row in raw_df.iterrows():
        if "Artículo" in row.values:
            header_row = i
            break
    
    if header_row is None:
        raise ValueError(f"No header found in {path}")
    
    # Starting header of the table
    try:
        df = pd.read_excel(path, header=(header_row+1))
    except Exception as e:
        # Use the same fallback method as above
        print(f"Warning: Standard Excel reading failed for data ({e}), using alternative method...")
        try:
            df = pd.read_excel(path, engine='xlrd', header=(header_row+1))
        except:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                data = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= header_row + 1:  # Skip to header row
                        data.append(row)
                df = pd.DataFrame(data)
                # Set column names from the header row
                header_data = list(ws.iter_rows(values_only=True))[header_row]
                df.columns = header_data[:len(df.columns)]
            except Exception as final_error:
                raise ValueError(f"Could not read Excel data from {path}: {final_error}")

    # Drop completely empty columns
    df.dropna(axis=1, inplace=True, how="all")
    print(df.head())
    df.drop(columns = ["U. med.", "Venta"], inplace=True)
    

    #Dropping last non table row from Excel
    df = df.iloc[:-1]

    # Get Number of Sucursal
    sucursal = gs.get_sucursal(path)

    df.columns = ["Código", "Artículo", "Unidades"]
    
    return df, sucursal
