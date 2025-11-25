"""
Excel Export Service for Statistics Data

This service generates professional Excel files from statistics data,
matching the format displayed in the Statistics frontend page.

Author: Embler Autopartes Analytics System
"""

import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

class StatisticsExcelExporter:
    """
    Professional Excel exporter for Statistics data
    
    Creates Excel files that match the Statistics page layout:
    - Product Code, Product Name columns
    - Dynamic store columns based on selected stores
    - Total Sold, Store Count, Average columns
    - Professional formatting with colors and borders
    """
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        
    def create_excel_from_statistics(
        self, 
        statistics_data: List[Dict[str, Any]], 
        stores_data: List[Dict[str, Any]],
        filters: Dict[str, Any] = None
    ) -> io.BytesIO:
        """
        Create Excel file from statistics data
        
        Args:
            statistics_data: List of product statistics (from API)
            stores_data: List of store information
            filters: Applied filters for documentation
            
        Returns:
            BytesIO: Excel file in memory
        """
        try:
            logger.info(f"Creating Excel export for {len(statistics_data)} products and {len(stores_data)} stores")
            
            # Store filters for use in table creation
            self._filters = filters
            
            # Debug: Log sample data structure
            if statistics_data:
                logger.info(f"Sample statistics data: {statistics_data[0]}")
            if stores_data:
                logger.info(f"Sample stores data: {stores_data[0]}")
            if filters and filters.get('advanced_filter'):
                logger.info(f"Advanced filter config: {filters['advanced_filter']}")
            
            # Create workbook and worksheet
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Estadísticas de Productos"
            
            # Add header information
            self._add_header_info(filters)
            
            # Create the main data table
            self._create_data_table(statistics_data, stores_data)
            
            # Apply professional formatting
            self._apply_formatting()
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            self.workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info("Excel file created successfully")
            return excel_buffer
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}", exc_info=True)
            raise
    
    def _add_header_info(self, filters: Dict[str, Any] = None):
        """Add header information to the Excel file"""
        
        # Title
        self.worksheet['A1'] = "Estadísticas de Productos - Embler Autopartes"
        self.worksheet['A1'].font = Font(size=16, bold=True, color="2F5597")
        
        # Generation date
        self.worksheet['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        self.worksheet['A2'].font = Font(size=10, italic=True)
        
        # Filter information
        if filters:
            row = 4
            self.worksheet[f'A{row}'] = "Filtros aplicados:"
            self.worksheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Check if advanced filter is used
            if filters.get('is_advanced_mode') and filters.get('advanced_filter'):
                self.worksheet[f'A{row}'] = "• Modo: Filtro Avanzado"
                self.worksheet[f'A{row}'].font = Font(color="2F5597")
                row += 1
                
                # Show advanced filter details
                advanced_config = filters['advanced_filter']
                if 'stores' in advanced_config:
                    stores_count = len(advanced_config['stores'])
                    self.worksheet[f'A{row}'] = f"• Sucursales configuradas: {stores_count}"
                    row += 1
                    
                    # Count sales and inventory selections
                    sales_count = sum(1 for config in advanced_config['stores'].values() if config.get('sales', False))
                    inventory_count = sum(1 for config in advanced_config['stores'].values() if config.get('inventory', False))
                    
                    if sales_count > 0:
                        self.worksheet[f'A{row}'] = f"• Columnas de ventas: {sales_count}"
                        row += 1
                    if inventory_count > 0:
                        self.worksheet[f'A{row}'] = f"• Columnas de inventario: {inventory_count}"
                        row += 1
            else:
                # Legacy mode filters
                if filters.get('type') and filters['type'] != 'all':
                    type_label = 'Ventas' if filters['type'] == 'sale' else 'Inventario'
                    self.worksheet[f'A{row}'] = f"• Tipo: {type_label}"
                    row += 1
                    
                if filters.get('selectedStores') and len(filters['selectedStores']) > 0:
                    self.worksheet[f'A{row}'] = f"• Sucursales seleccionadas: {len(filters['selectedStores'])}"
                    row += 1
                
            # Date filters (common to both modes)
            if filters.get('date_from') and filters.get('date_to'):
                self.worksheet[f'A{row}'] = f"• Fecha: {filters['date_from']} a {filters['date_to']}"
                row += 1
            elif filters.get('date_from'):
                self.worksheet[f'A{row}'] = f"• Fecha desde: {filters['date_from']}"
                row += 1
            elif filters.get('date_to'):
                self.worksheet[f'A{row}'] = f"• Fecha hasta: {filters['date_to']}"
                row += 1
            elif filters.get('date'):
                self.worksheet[f'A{row}'] = f"• Fecha: {filters['date']}"
                row += 1
        
        # Empty row before table
        self._table_start_row = max(7, row + 2) if filters else 5
    
    def _create_data_table(self, statistics_data: List[Dict[str, Any]], stores_data: List[Dict[str, Any]]):
        """Create the enhanced data table with support for both sales and inventory"""
        
        try:
            logger.info(f"Creating enhanced data table with {len(statistics_data)} products and {len(stores_data)} stores")
            
            # Create store lookup
            store_map = {store['id']: store['name'] for store in stores_data}
            logger.info(f"Store map created: {store_map}")
            
            # Get all store IDs that appear in the data (both sales and inventory)
            all_store_ids = set()
            has_inventory_data = False
            
            for product in statistics_data:
                store_sales = product.get('store_sales', {})
                store_inventory = product.get('store_inventory', {})
                
                all_store_ids.update(store_sales.keys())
                all_store_ids.update(store_inventory.keys())
                
                if store_inventory:
                    has_inventory_data = True
            
            # Convert string keys to integers for consistent sorting
            all_store_ids = {int(sid) if isinstance(sid, str) else sid for sid in all_store_ids}
            sorted_store_ids = sorted(all_store_ids)
            
            logger.info(f"All store IDs found: {sorted_store_ids}, has_inventory_data: {has_inventory_data}")
            
        except Exception as e:
            logger.error(f"Error in _create_data_table setup: {str(e)}", exc_info=True)
            raise
        
        # Create headers based on data structure
        headers = ['Código', 'Producto']
        
        # Determine column structure based on advanced filter configuration
        if has_inventory_data:
            # Advanced mode: use filter configuration to determine columns
            # Check if we have advanced filter configuration in filters
            advanced_config = None
            if self._filters and self._filters.get('advanced_filter'):
                advanced_config = self._filters['advanced_filter']
            
            if advanced_config and 'stores' in advanced_config:
                # Use advanced filter configuration to determine columns
                for store_id_str, store_config in advanced_config['stores'].items():
                    store_id = int(store_id_str)
                    store_name = store_map.get(store_id, f'Sucursal {store_id}')
                    
                    show_sales = store_config.get('sales', False)
                    show_inventory = store_config.get('inventory', False)
                    
                    if show_sales and show_inventory:
                        headers.extend([f'{store_name} (V)', f'{store_name} (I)'])
                    elif show_sales:
                        headers.append(f'{store_name} (V)')
                    elif show_inventory:
                        headers.append(f'{store_name} (I)')
            else:
                # Fallback: check data for column structure
                for store_id in sorted_store_ids:
                    store_name = store_map.get(store_id, f'Sucursal {store_id}')
                    
                    # Check if this store has both sales and inventory data
                    has_sales = any(product.get('store_sales', {}).get(store_id, 0) > 0 for product in statistics_data)
                    has_inventory = any(product.get('store_inventory', {}).get(store_id, 0) > 0 for product in statistics_data)
                    
                    if has_sales and has_inventory:
                        headers.extend([f'{store_name} (V)', f'{store_name} (I)'])
                    elif has_sales:
                        headers.append(f'{store_name} (V)')
                    elif has_inventory:
                        headers.append(f'{store_name} (I)')
            
            # Enhanced summary columns
            headers.extend(['Total Ventas', 'Total Inventario', 'Sucursales', 'Promedio Ventas', 'Promedio Inventario'])
        else:
            # Legacy mode: single column per store
            for store_id in sorted_store_ids:
                store_name = store_map.get(store_id, f'Sucursal {store_id}')
                headers.append(store_name)
            
            headers.extend(['Total Vendido', 'Sucursales', 'Promedio'])
        
        # Write headers
        header_row = self._table_start_row
        for col_idx, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            
            # Color code headers
            if '(V)' in header or 'Ventas' in header:
                cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")  # Green for sales
            elif '(I)' in header or 'Inventario' in header:
                cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Blue for inventory
            else:
                cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")  # Default blue
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data rows
        for row_idx, product in enumerate(statistics_data, header_row + 1):
            col_idx = 1
            
            # Product code
            self.worksheet.cell(row=row_idx, column=col_idx, value=product['product_code'])
            col_idx += 1
            
            # Product name
            self.worksheet.cell(row=row_idx, column=col_idx, value=product['product_name'])
            col_idx += 1
            
            # Store data
            store_sales = product.get('store_sales', {})
            store_inventory = product.get('store_inventory', {})
            
            total_sales = 0
            total_inventory = 0
            sales_store_count = 0
            inventory_store_count = 0
            
            if has_inventory_data:
                # Advanced mode: handle both sales and inventory
                advanced_config = None
                if self._filters and self._filters.get('advanced_filter'):
                    advanced_config = self._filters['advanced_filter']
                
                if advanced_config and 'stores' in advanced_config:
                    # Use advanced filter configuration for data columns
                    for store_id_str, store_config in advanced_config['stores'].items():
                        store_id = int(store_id_str)
                        
                        # Get sales and inventory quantities
                        sales_qty = store_sales.get(store_id, store_sales.get(str(store_id), 0)) or 0
                        inventory_qty = store_inventory.get(store_id, store_inventory.get(str(store_id), 0)) or 0
                        
                        show_sales = store_config.get('sales', False)
                        show_inventory = store_config.get('inventory', False)
                        
                        if show_sales and show_inventory:
                            # Both columns
                            self.worksheet.cell(row=row_idx, column=col_idx, value=sales_qty)
                            col_idx += 1
                            self.worksheet.cell(row=row_idx, column=col_idx, value=inventory_qty)
                            col_idx += 1
                        elif show_sales:
                            # Sales only
                            self.worksheet.cell(row=row_idx, column=col_idx, value=sales_qty)
                            col_idx += 1
                        elif show_inventory:
                            # Inventory only
                            self.worksheet.cell(row=row_idx, column=col_idx, value=inventory_qty)
                            col_idx += 1
                        
                        # Update totals
                        total_sales += sales_qty
                        total_inventory += inventory_qty
                        if sales_qty > 0:
                            sales_store_count += 1
                        if inventory_qty > 0:
                            inventory_store_count += 1
                else:
                    # Fallback: use data-based logic
                    for store_id in sorted_store_ids:
                        # Get sales and inventory quantities
                        sales_qty = store_sales.get(store_id, store_sales.get(str(store_id), 0)) or 0
                        inventory_qty = store_inventory.get(store_id, store_inventory.get(str(store_id), 0)) or 0
                        
                        # Check if this store has both types of data
                        has_sales = any(product.get('store_sales', {}).get(store_id, 0) > 0 for product in statistics_data)
                        has_inventory_for_store = any(product.get('store_inventory', {}).get(store_id, 0) > 0 for product in statistics_data)
                        
                        if has_sales and has_inventory_for_store:
                            # Both columns
                            self.worksheet.cell(row=row_idx, column=col_idx, value=sales_qty)
                            col_idx += 1
                            self.worksheet.cell(row=row_idx, column=col_idx, value=inventory_qty)
                            col_idx += 1
                        elif has_sales:
                            # Sales only
                            self.worksheet.cell(row=row_idx, column=col_idx, value=sales_qty)
                            col_idx += 1
                        elif has_inventory_for_store:
                            # Inventory only
                            self.worksheet.cell(row=row_idx, column=col_idx, value=inventory_qty)
                            col_idx += 1
                        
                        # Update totals
                        total_sales += sales_qty
                        total_inventory += inventory_qty
                        if sales_qty > 0:
                            sales_store_count += 1
                        if inventory_qty > 0:
                            inventory_store_count += 1
                
                # Enhanced summary columns
                self.worksheet.cell(row=row_idx, column=col_idx, value=total_sales)
                col_idx += 1
                self.worksheet.cell(row=row_idx, column=col_idx, value=total_inventory)
                col_idx += 1
                
                # Combined store count
                combined_store_count = len(set(
                    [sid for sid in store_sales.keys() if store_sales[sid] > 0] +
                    [sid for sid in store_inventory.keys() if store_inventory[sid] > 0]
                ))
                self.worksheet.cell(row=row_idx, column=col_idx, value=combined_store_count)
                col_idx += 1
                
                # Sales average
                sales_avg = total_sales / sales_store_count if sales_store_count > 0 else 0
                self.worksheet.cell(row=row_idx, column=col_idx, value=round(sales_avg, 2))
                col_idx += 1
                
                # Inventory average
                inventory_avg = total_inventory / inventory_store_count if inventory_store_count > 0 else 0
                self.worksheet.cell(row=row_idx, column=col_idx, value=round(inventory_avg, 2))
                
            else:
                # Legacy mode: single column per store (sales only)
                for store_id in sorted_store_ids:
                    quantity = store_sales.get(store_id, store_sales.get(str(store_id), 0)) or 0
                    self.worksheet.cell(row=row_idx, column=col_idx, value=quantity)
                    total_sales += quantity
                    if quantity > 0:
                        sales_store_count += 1
                    col_idx += 1
                
                # Legacy summary columns
                self.worksheet.cell(row=row_idx, column=col_idx, value=total_sales)
                col_idx += 1
                self.worksheet.cell(row=row_idx, column=col_idx, value=sales_store_count)
                col_idx += 1
                
                average = total_sales / sales_store_count if sales_store_count > 0 else 0
                self.worksheet.cell(row=row_idx, column=col_idx, value=round(average, 2))
        
        self._table_end_row = header_row + len(statistics_data)
        self._table_end_col = len(headers)
    
    def _apply_formatting(self):
        """Apply professional formatting to the Excel file"""
        
        # Auto-adjust column widths
        for col_idx in range(1, self._table_end_col + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Iterate through all rows in this column
            for row_idx in range(1, self._table_end_row + 1):
                cell = self.worksheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply borders to data table
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(self._table_start_row, self._table_end_row + 1):
            for col in range(1, self._table_end_col + 1):
                self.worksheet.cell(row=row, column=col).border = thin_border
        
        # Alternate row colors for better readability
        light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row in range(self._table_start_row + 1, self._table_end_row + 1):
            if (row - self._table_start_row) % 2 == 0:  # Even rows
                for col in range(1, self._table_end_col + 1):
                    self.worksheet.cell(row=row, column=col).fill = light_fill
        
        # Center align numeric columns
        for row in range(self._table_start_row, self._table_end_row + 1):
            for col in range(3, self._table_end_col + 1):  # Skip product code and name
                cell = self.worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Freeze panes (freeze header row and first two columns)
        self.worksheet.freeze_panes = f'C{self._table_start_row + 1}'


def create_statistics_excel_export(
    statistics_data: List[Dict[str, Any]], 
    stores_data: List[Dict[str, Any]],
    filters: Dict[str, Any] = None
) -> io.BytesIO:
    """
    Convenience function to create Excel export
    
    Args:
        statistics_data: Product statistics data
        stores_data: Store information
        filters: Applied filters for documentation
        
    Returns:
        BytesIO: Excel file ready for download
    """
    exporter = StatisticsExcelExporter()
    return exporter.create_excel_from_statistics(statistics_data, stores_data, filters)
